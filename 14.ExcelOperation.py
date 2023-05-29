
import openpyxl

filename = "F:\\Testing\\Database_Testing\\Customer.xlsx"
# Load /open excel file by name
workbook = openpyxl.load_workbook(filename)

# Locate /get particular sheet
sheet = workbook['Sheet1']
sheet_obj = sheet.active
m_row = sheet_obj.max_row

for i in range (1, m_row+1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    print(cell_obj.value)

